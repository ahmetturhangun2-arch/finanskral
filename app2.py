from flask import Flask, render_template, request, Response, redirect, url_for, send_file, flash, jsonify

import csv
import io
import os
import tempfile
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from werkzeug.utils import secure_filename

app = Flask(__name__)

ADSENSE_ENABLED = False
ADSENSE_CLIENT = "ca-pub-8518693068868746"
ADSENSE_TOP_SLOT = "1111111111"
ADSENSE_CONTENT_SLOT = "2222222222"
ADSENSE_INLINE_SLOT = "3333333333"
ADSENSE_BOTTOM_SLOT = "4444444444"
ADSENSE_LEFT_SLOT = "5555555555"
ADSENSE_RIGHT_SLOT = "6666666666"


AVERAGE_MONTHLY_CAR_COST = 6500
AVERAGE_SAVINGS_RATE = 20

NEWS_CACHE = {"items": [], "timestamp": 0}
NEWS_CACHE_TTL = 900

CRYPTO_CACHE = {"items": [], "timestamp": 0}
CRYPTO_CACHE_TTL = 120
CRYPTO_FALLBACK = [
    {"symbol": "BTC", "name": "Bitcoin", "price": 0, "change_24h": 0},
    {"symbol": "ETH", "name": "Ethereum", "price": 0, "change_24h": 0},
    {"symbol": "BNB", "name": "BNB", "price": 0, "change_24h": 0},
    {"symbol": "SOL", "name": "Solana", "price": 0, "change_24h": 0},
    {"symbol": "XRP", "name": "XRP", "price": 0, "change_24h": 0},
    {"symbol": "DOGE", "name": "Dogecoin", "price": 0, "change_24h": 0},
]


ECONOMY_HEADLINES = [
    {
        "tag": "Piyasa",
        "title": "Dolar ve Eurodaki Son Hareketler Maaş Alım Gücünü Nasıl Etkiler?",
        "desc": "Kur değişimleri, market, kira ve araç giderleri üzerinde doğrudan baskı oluşturabilir.",
        "url": "https://news.google.com/search?q=dolar%20euro%20tl%20ekonomi%20haberleri&hl=tr&gl=TR&ceid=TR:tr",
    },
    {
        "tag": "Akaryakıt",
        "title": "Benzin ve Motorin Fiyatları Araç Maliyetini Yeniden Şekillendiriyor",
        "desc": "Yakıt fiyatlarındaki değişim, aylık araç gideri ve kilometre başı maliyet hesabını etkiler.",
        "url": "https://news.google.com/search?q=akaryakit%20fiyatlari%20benzin%20motorin%20haberleri&hl=tr&gl=TR&ceid=TR:tr",
    },
    {
        "tag": "Bütçe",
        "title": "Kira, Market ve Fatura Giderleri Hane Bütçesinde Öne Çıkıyor",
        "desc": "Temel giderlerdeki artış, maaş gelir-gider analizini daha önemli hale getiriyor.",
        "url": "https://news.google.com/search?q=kira%20market%20fatura%20giderleri%20ekonomi&hl=tr&gl=TR&ceid=TR:tr",
    },
    {
        "tag": "Otomobil",
        "title": "Araç Sahibi Olmanın Gerçek Aylık Maliyeti Neden Artıyor?",
        "desc": "Sigorta, bakım, MTV ve yakıt kalemleri toplam araç maliyetinin ana parçalarını oluşturur.",
        "url": "https://news.google.com/search?q=arac%20maliyeti%20sigorta%20mtv%20bakim%20haberleri&hl=tr&gl=TR&ceid=TR:tr",
    },
    {
        "tag": "Tasarruf",
        "title": "Gelir-Gider Dengesi İçin En Çok Takip Edilen Finans Başlıkları",
        "desc": "Maaş, borç, kart ekstresi ve tasarruf alışkanlıkları aylık finans planının merkezinde yer alır.",
        "url": "https://news.google.com/search?q=tasarruf%20gelir%20gider%20butce%20haberleri&hl=tr&gl=TR&ceid=TR:tr",
    },
    {
        "tag": "Rehber",
        "title": "Maaş Yetmiyorsa İlk Bakılması Gereken 5 Gider Kalemi",
        "desc": "Büyük giderleri görünür hale getirmek, bütçeyi toparlamanın ilk adımıdır.",
        "url": "https://news.google.com/search?q=maas%20butce%20gider%20yonetimi%20ekonomi&hl=tr&gl=TR&ceid=TR:tr",
    },
]

CITY_LIVING_COSTS = [
    {"city": "İstanbul", "student": "28.000 - 42.000 TL", "single": "38.000 - 58.000 TL", "car": "8.500 - 13.500 TL", "note": "Kira ve ulaşım yükü yüksek."},
    {"city": "Ankara", "student": "22.000 - 34.000 TL", "single": "31.000 - 47.000 TL", "car": "7.500 - 12.000 TL", "note": "Kira İstanbul'a göre daha dengeli."},
    {"city": "İzmir", "student": "24.000 - 37.000 TL", "single": "33.000 - 50.000 TL", "car": "7.800 - 12.500 TL", "note": "Kıyı ilçelerinde kira farkı artabilir."},
    {"city": "Adana", "student": "19.000 - 30.000 TL", "single": "27.000 - 41.000 TL", "car": "7.000 - 11.500 TL", "note": "Araç kullanımı ve yakıt bütçesi öne çıkar."},
    {"city": "Bursa", "student": "21.000 - 33.000 TL", "single": "30.000 - 45.000 TL", "car": "7.400 - 12.000 TL", "note": "Sanayi ve ulaşım maliyetleri belirleyici."},
    {"city": "Antalya", "student": "23.000 - 36.000 TL", "single": "32.000 - 49.000 TL", "car": "7.800 - 12.800 TL", "note": "Sezonluk fiyat değişimi güçlü."},
]

DAILY_ECONOMY_IMPACTS = [
    {"label": "Yakıt Etkisi", "title": "Motorine gelen her 1 TL artış, 1000 km/ay kullanımda bütçeyi hissedilir yükseltir.", "metric": "Araç sahipleri"},
    {"label": "Kur Etkisi", "title": "Dolar/TL hareketleri elektronik, otomobil ve ithal ürün maliyetini etkileyebilir.", "metric": "Alım gücü"},
    {"label": "Market Etkisi", "title": "Market sepetindeki küçük artışlar ay sonunda en büyük görünmeyen gider olabilir.", "metric": "Günlük bütçe"},
    {"label": "Kredi Etkisi", "title": "Kart borcu ve gecikme faizleri maaşın daha hesaba yatmadan erimesine yol açabilir.", "metric": "Borç riski"},
]

FINANCIAL_DEATH_ZONES = [
    {"title": "Maaşı bitiren 5 alışkanlık", "desc": "Küçük abonelikler, plansız alışveriş ve günlük kart harcamaları ay sonunda büyük fark yaratır."},
    {"title": "Araba sahiplerini fakirleştiren gizli gider", "desc": "Yakıt dışındaki sigorta, bakım, MTV, lastik ve otopark kalemleri gerçek maliyeti büyütür."},
    {"title": "Kredi kartı limiti neden tehlikeli?", "desc": "Limit gelir gibi hissedildiğinde gelecek ayın maaşı bugünden harcanır."},
    {"title": "Asgari ücretlinin görünmeyen vergileri", "desc": "Dolaylı vergiler, yakıt ve tüketim harcamaları üzerinden bütçeye sürekli yansır."},
]

BLOG_POSTS = [
    {
        "slug": "kredi-notu-nasil-yukseltilir",
        "title": "Kredi Notu Nasıl Yükseltilir?",
        "description": "Kredi notunu etkileyen temel faktörler ve düzenli ödeme alışkanlığıyla puanı iyileştirme rehberi.",
        "content": [
            "Kredi notu, bankaların kişinin ödeme alışkanlığını ve finansal disiplinini anlamak için kullandığı önemli göstergelerden biridir. Düzenli ödeme, düşük borçluluk oranı ve aktif finansal geçmiş bu puanın güçlenmesine yardımcı olabilir.",
            "İlk adım, kredi kartı ve kredi taksitlerini son ödeme tarihinden önce ödemektir. Gecikmeler kısa vadede küçük görünse bile finansal geçmişte olumsuz iz bırakabilir.",
            "Kredi kartı limitinin tamamını sürekli kullanmak da riskli algılanabilir. Kullanım oranını makul seviyede tutmak ve borcu her ay planlı kapatmak daha sağlıklı bir görünüm oluşturur.",
            "Yeni kredi başvurularını çok sık yapmak yerine, mevcut borçları düzenlemek ve gelir-gider dengesini korumak uzun vadede daha güçlü bir kredi profili oluşturur."
        ]
    },
    {
        "slug": "aylik-butce-nasil-yapilir",
        "title": "Aylık Bütçe Nasıl Yapılır?",
        "description": "Gelir, sabit gider, değişken gider ve birikim hedeflerini tek planda toplama yöntemi.",
        "content": [
            "Aylık bütçe, paranın nereye gittiğini görmek ve ay sonunda sürpriz yaşamamak için hazırlanır. En basit yöntem, geliri ve tüm giderleri ayrı başlıklar halinde yazmaktır.",
            "Kira, fatura, ulaşım ve market gibi düzenli giderler sabit kalemler olarak ayrılmalıdır. Eğlence, alışveriş ve plansız harcamalar ise değişken gider olarak takip edilebilir.",
            "Birikim bütçenin sonunda kalan para değil, baştan ayrılan bir hedef olmalıdır. Maaş gelir gelmez küçük de olsa belirli bir tutarı farklı hesaba aktarmak disiplini artırır.",
            "Bütçe ayda bir kez kontrol edilmeli, gereksiz harcamalar tespit edilmeli ve sonraki ay için gerçekçi limitler belirlenmelidir."
        ]
    },
    {
        "slug": "arac-maliyeti-hesaplama-rehberi",
        "title": "Araç Maliyeti Hesaplama Rehberi",
        "description": "Yakıt, bakım, sigorta ve MTV kalemleriyle aracın gerçek aylık maliyetini hesaplama.",
        "content": [
            "Araç sahibi olmak yalnızca yakıt masrafından ibaret değildir. Sigorta, bakım, MTV, lastik, muayene ve beklenmeyen arızalar toplam maliyeti belirler.",
            "Aylık araç maliyeti hesaplanırken yıllık giderler 12 aya bölünmeli ve yakıt tüketimiyle birlikte değerlendirilmelidir. Böylece aracın bütçeye gerçek etkisi daha net görülür.",
            "Kilometre başına maliyet, farklı araçları karşılaştırmak için kullanışlıdır. Aynı mesafeyi daha düşük tüketimle yapan araç uzun vadede ciddi tasarruf sağlayabilir.",
            "Araç giderlerini azaltmak için düzenli bakım, doğru lastik basıncı, sakin sürüş ve sigorta tekliflerini karşılaştırmak etkili yöntemlerdir."
        ]
    },
    {
        "slug": "maas-yonetimi-icin-pratik-oneriler",
        "title": "Maaş Yönetimi İçin Pratik Öneriler",
        "description": "Maaşı daha verimli kullanmak, giderleri dengelemek ve birikim alışkanlığı kazanmak için öneriler.",
        "content": [
            "Maaş yönetimi, gelirin büyüklüğünden çok paranın nasıl dağıtıldığıyla ilgilidir. Düzenli plan yapılmadığında yüksek gelir bile ay sonunda yetersiz kalabilir.",
            "Maaş geldiği gün kira, fatura, kart ödemesi ve birikim gibi temel kalemler ayrılmalıdır. Kalan tutar günlük harcama limiti olarak düşünülmelidir.",
            "Küçük harcamalar takip edilmediğinde bütçeyi zorlayabilir. Kahve, abonelik, uygulama ve plansız alışverişler ay sonunda büyük toplam oluşturabilir.",
            "FinansKral gibi hesaplama araçları, maaş ve gider dengesini hızlı görmeye yardımcı olur. Düzenli takip, daha sağlıklı kararların temelidir."
        ]
    },
    {
        "slug": "acil-durum-fonu-nedir",
        "title": "Acil Durum Fonu Nedir?",
        "description": "Beklenmeyen masraflara karşı güvenli nakit alanı oluşturmanın önemi.",
        "content": [
            "Acil durum fonu, iş kaybı, sağlık masrafı, araç arızası veya beklenmeyen giderler için ayrılan güvenli paradır. Bu fon yatırım amacıyla değil, finansal güvenlik amacıyla tutulur.",
            "Genellikle birkaç aylık temel gideri karşılayabilecek tutar hedeflenir. Ancak başlangıç için küçük ve düzenli birikim yapmak da değerlidir.",
            "Bu para kolay ulaşılabilir fakat günlük harcamalarla karışmayacak ayrı bir hesapta tutulmalıdır. Böylece ihtiyaç anında kredi kartına yüklenme ihtiyacı azalır.",
            "Acil durum fonu tamamlandıktan sonra yatırım, borç kapatma veya daha büyük hedefler için plan yapmak daha sağlıklı olur."
        ]
    },
    {
        "slug": "market-harcamalari-nasil-azaltilir",
        "title": "Market Harcamaları Nasıl Azaltılır?",
        "description": "Liste, haftalık plan ve fiyat karşılaştırmasıyla market bütçesini kontrol etme yolları.",
        "content": [
            "Market harcamaları çoğu bütçede en hızlı büyüyen kalemlerden biridir. Plansız alışveriş, küçük görünen ürünlerin toplamını ciddi seviyeye çıkarabilir.",
            "Alışverişe listeyle gitmek ve haftalık yemek planı yapmak gereksiz ürün alımını azaltır. Açken alışveriş yapmamak da basit ama etkili bir yöntemdir.",
            "Birim fiyat karşılaştırması yapmak, büyük ambalajın gerçekten avantajlı olup olmadığını gösterir. Kampanya ürünlerinde ihtiyaç dışı alım yapılmamalıdır.",
            "Market bütçesi için haftalık limit belirlemek ve harcamaları düzenli kaydetmek, ay sonundaki sapmaları azaltır."
        ]
    },
    {
        "slug": "yakittan-tasarruf-etme-yollari",
        "title": "Yakıttan Tasarruf Etme Yolları",
        "description": "Sürüş alışkanlığı ve bakım düzeniyle yakıt giderlerini azaltma yöntemleri.",
        "content": [
            "Yakıt tasarrufu yalnızca araç modeliyle değil, sürüş tarzıyla da doğrudan ilişkilidir. Ani hızlanma ve sert fren tüketimi artırabilir.",
            "Düzenli bakım, temiz hava filtresi ve doğru lastik basıncı tüketimi olumlu etkileyebilir. Küçük ihmal edilen bakım kalemleri uzun vadede daha yüksek yakıt gideri oluşturur.",
            "Rota planı yapmak, yoğun trafikten kaçınmak ve gereksiz yük taşımamak tasarrufa katkı sağlar. Kısa mesafelerde alternatif ulaşım düşünmek de bütçeyi rahatlatabilir.",
            "Aylık kilometre ve yakıt harcamasını takip etmek, aracın gerçek tüketimini görmenin en pratik yoludur."
        ]
    },
    {
        "slug": "borc-kapatma-plani-nasil-yapilir",
        "title": "Borç Kapatma Planı Nasıl Yapılır?",
        "description": "Borçları listeleyip önceliklendirme ve düzenli ödeme planı oluşturma rehberi.",
        "content": [
            "Borç kapatma planı, tüm borçları görünür hale getirmekle başlar. Kredi, kredi kartı, taksit ve kişisel borçlar ayrı ayrı yazılmalıdır.",
            "Her borç için kalan tutar, faiz oranı ve son ödeme tarihi takip edilmelidir. Yüksek faizli borçlara öncelik vermek toplam maliyeti azaltabilir.",
            "Asgari ödeme döngüsünde kalmak uzun vadede maliyeti artırabilir. Mümkünse bütçede ek ödeme alanı açmak borcun daha hızlı kapanmasını sağlar.",
            "Yeni borç almadan önce mevcut borçların sürdürülebilir olup olmadığı hesaplanmalı ve aylık nakit akışı korunmalıdır."
        ]
    },
]


def _strip_source_from_title(title, source):
    title = title or "Ekonomi haberi"
    if source and title.endswith(f" - {source}"):
        title = title[:-(len(source) + 3)].strip()
    return title


def _category_for_news(title):
    text = (title or "").lower()
    if any(word in text for word in ["benzin", "motorin", "akaryak", "yakıt", "yakit", "lpg"]):
        return "Akaryakıt"
    if any(word in text for word in ["dolar", "euro", "kur", "altın", "altin"]):
        return "Piyasa"
    if any(word in text for word in ["enflasyon", "zam", "fiyat", "market"]):
        return "Enflasyon"
    if any(word in text for word in ["maaş", "maas", "asgari", "emekli", "memur"]):
        return "Gelir"
    if any(word in text for word in ["otomobil", "araç", "arac", "mtv", "sigorta"]):
        return "Araç"
    return "Ekonomi"


def get_google_news_items(limit=18):
    now = time.time()
    cached = NEWS_CACHE.get("items", [])
    if cached and now - NEWS_CACHE.get("timestamp", 0) < NEWS_CACHE_TTL:
        return cached[:limit]

    query = 'ekonomi OR dolar OR euro OR akaryakıt OR benzin OR motorin OR enflasyon OR maaş OR asgari ücret'
    url = (
        "https://news.google.com/rss/search?q="
        + urllib.parse.quote(query)
        + "&hl=tr&gl=TR&ceid=TR:tr"
    )

    items = []
    try:
        request_obj = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 FinansKral/1.0 (+https://finanskral.com)"
            },
        )
        with urllib.request.urlopen(request_obj, timeout=6) as response:
            xml_data = response.read()

        root = ET.fromstring(xml_data)
        for item in root.findall("./channel/item"):
            title_raw = (item.findtext("title") or "").strip()
            link = (item.findtext("link") or "").strip()
            pub_date = (item.findtext("pubDate") or "").strip()
            source_el = item.find("source")
            source = (source_el.text or "Google News") if source_el is not None else "Google News"
            source_url = source_el.attrib.get("url", "") if source_el is not None else ""
            domain = urllib.parse.urlparse(source_url).netloc.replace("www.", "") if source_url else "news.google.com"
            title = _strip_source_from_title(title_raw, source)
            if not title:
                continue
            items.append({
                "title": title,
                "source": source,
                "source_domain": domain,
                "link": link,
                "pub_date": pub_date,
                "category": _category_for_news(title),
                "favicon": f"https://www.google.com/s2/favicons?domain={domain}&sz=64",
            })
            if len(items) >= limit:
                break

        if items:
            NEWS_CACHE["items"] = items
            NEWS_CACHE["timestamp"] = now
            return items[:limit]
    except Exception:
        pass

    fallback = []
    for item in ECONOMY_HEADLINES:
        fallback.append({
            "title": item["title"],
            "source": "FinansKral",
            "source_domain": "finanskral.com",
            "link": item["url"],
            "pub_date": "",
            "category": item["tag"],
            "favicon": "",
        })
    return fallback[:limit]


def get_crypto_prices(limit=6):
    now = time.time()
    if CRYPTO_CACHE["items"] and now - CRYPTO_CACHE["timestamp"] < CRYPTO_CACHE_TTL:
        return CRYPTO_CACHE["items"][:limit]

    url = (
        "https://api.coingecko.com/api/v3/coins/markets?"
        "vs_currency=usd&ids=bitcoin,ethereum,binancecoin,solana,ripple,dogecoin"
        "&order=market_cap_desc&per_page=6&page=1&sparkline=false"
        "&price_change_percentage=24h"
    )
    symbol_map = {
        "bitcoin": "BTC",
        "ethereum": "ETH",
        "binancecoin": "BNB",
        "solana": "SOL",
        "ripple": "XRP",
        "dogecoin": "DOGE",
    }
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "FinansKral/1.0"})
        with urllib.request.urlopen(req, timeout=6) as response:
            data = response.read().decode("utf-8")
        import json
        payload = json.loads(data)
        items = []
        for coin in payload:
            items.append({
                "symbol": symbol_map.get(coin.get("id"), (coin.get("symbol") or "").upper()),
                "name": coin.get("name", "Kripto"),
                "price": float(coin.get("current_price") or 0),
                "change_24h": float(coin.get("price_change_percentage_24h") or 0),
            })
        if items:
            CRYPTO_CACHE["items"] = items
            CRYPTO_CACHE["timestamp"] = now
            return items[:limit]
    except Exception:
        pass
    return CRYPTO_CACHE["items"][:limit] if CRYPTO_CACHE["items"] else CRYPTO_FALLBACK[:limit]


def format_tl(value):
    formatted = f"{value:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{formatted} TL"


def clamp(value, min_value, max_value):
    return max(min_value, min(value, max_value))


def safe_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def build_salary_tips(savings_rate, biggest_expense_label, ay_sonu_kalan):
    tips = []
    if savings_rate < 10:
        tips.append("Gelirinin en az %10'unu kenara koyacak otomatik bir birikim hedefi belirle.")
    if biggest_expense_label == "Kira":
        tips.append("Kira bütçeni gözden geçir; toplam gelirin %30–35 bandı daha sürdürülebilir olur.")
    elif biggest_expense_label == "Market":
        tips.append("Market giderleri için haftalık limit belirleyip toplu alışverişe geçmeyi dene.")
    elif biggest_expense_label == "Fatura":
        tips.append("Fatura kalemlerini tek tek inceleyip kullanmadığın abonelikleri kapat.")
    elif biggest_expense_label == "Diğer":
        tips.append("‘Diğer giderler’ kalemini parçalayarak hangi harcamanın bütçeyi deldiğini tespit et.")
    if ay_sonu_kalan < 0:
        tips.append("Ay sonu eksiye düştüğün için önce büyük gider kalemini küçült, sonra ek gelir alanı aç.")
    else:
        tips.append("Pozitif kalan tutarı aylık yatırım veya acil durum fonu için ayrı bir hesaba aktar.")
    return tips[:3]


def build_car_tips(km_basi_maliyet, yakit_share_pct, arac_skor):
    tips = []
    if yakit_share_pct >= 55:
        tips.append("Yakıt gideri toplam maliyetin büyük kısmını oluşturuyor; sürüş stilini ve rota planını optimize et.")
    if km_basi_maliyet > 7:
        tips.append("KM başı maliyet yüksek; daha ekonomik lastik, bakım planı veya alternatif araç tipi değerlendirmesi yap.")
    if arac_skor < 40:
        tips.append("Bu araç kullanım profili pahalı görünüyor; daha düşük tüketimli araç seçeneği uzun vadede ciddi fark yaratabilir.")
    if len(tips) < 3:
        tips.append("Sigorta ve bakım kalemlerini yılda en az bir kez karşılaştırarak toplam maliyeti aşağı çekebilirsin.")
    return tips[:3]



def build_viral_salary_lines(net_maas, toplam_gider, ay_sonu_kalan, biggest_expense_label):
    spent_pct = (toplam_gider / net_maas * 100) if net_maas > 0 else 0
    iphone_equiv = round(abs(toplam_gider * 12) / 75000, 1) if toplam_gider else 0
    if ay_sonu_kalan < 0:
        headline = "Bu ay bütçe eksiye düştü"
    elif spent_pct > 70:
        headline = "Maaşın büyük kısmı giderlere gidiyor"
    else:
        headline = "Bütçede nefes alanı var"
    return {
        "headline": headline,
        "line1": f"Maaşının %{round(spent_pct, 1)} kadarı giderlere gidiyor.",
        "line2": f"En büyük baskı noktası: {biggest_expense_label}.",
        "line3": f"Yıllık gider etkisi yaklaşık {iphone_equiv} iPhone bütçesine denk.",
    }


def build_car_affordability(monthly_income, aylik_toplam):
    if monthly_income <= 0:
        return {
            "ratio": "Gelir girilmedi",
            "verdict": "Gelirini eklersen aracın sana ağır gelip gelmediğini de hesaplarız.",
            "level": "neutral",
        }
    ratio = aylik_toplam / monthly_income * 100 if monthly_income else 0
    if ratio <= 15:
        verdict = "Bu araç gelirine göre rahat bölgede görünüyor."
        level = "good"
    elif ratio <= 30:
        verdict = "Bu araç gelirine göre dikkat gerektiren bölgede."
        level = "mid"
    else:
        verdict = "Bu araç gelirine göre ağır maliyet yaratıyor."
        level = "risk"
    return {"ratio": f"%{round(ratio, 1)}", "verdict": verdict, "level": level}


def build_viral_car_lines(aylik_toplam, yearly_total, km_basi_maliyet, affordability):
    return {
        "headline": "Arabanın gerçek maliyeti ortaya çıktı",
        "line1": f"Aylık araç yükü: {format_tl(aylik_toplam)}.",
        "line2": f"Yıllık toplam etki: {format_tl(yearly_total)}.",
        "line3": f"KM başına maliyet: {format_tl(km_basi_maliyet)}. {affordability['verdict']}",
    }

def save_lead(payload):
    data_dir = Path(__file__).resolve().parent / "data"
    data_dir.mkdir(exist_ok=True)
    csv_path = data_dir / "lead_requests.csv"
    fieldnames = [
        "timestamp", "source", "name", "email", "phone", "goal", "summary", "page"
    ]
    file_exists = csv_path.exists()
    with csv_path.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow({key: payload.get(key, "") for key in fieldnames})



@app.route("/api/crypto-prices")
def api_crypto_prices():
    return jsonify({"items": get_crypto_prices(limit=6)})

@app.route("/robots.txt")
def robots_txt():
    content = """User-agent: *
Allow: /

Sitemap: /sitemap.xml
"""
    return Response(content, mimetype="text/plain")


@app.route("/sitemap.xml")
def sitemap_xml():
    base_url = "https://finanskral.com"
    urls = [
        ("/", "weekly", "1.0"),
        ("/about", "monthly", "0.7"),
        ("/contact", "monthly", "0.6"),
        ("/privacy", "yearly", "0.5"),
        ("/cookies", "yearly", "0.5"),
        ("/terms", "yearly", "0.5"),
        ("/blog", "weekly", "0.8"),
        ("/haberler", "hourly", "0.9"),
    ]
    urls.extend((f"/blog/{post['slug']}", "monthly", "0.7") for post in BLOG_POSTS)
    items = "\n".join(
        f"  <url><loc>{base_url}{path}</loc><changefreq>{freq}</changefreq><priority>{priority}</priority></url>"
        for path, freq, priority in urls
    )
    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
{items}
</urlset>
"""
    return Response(xml, mimetype="application/xml")

@app.route("/", methods=["GET", "POST"])
def home():
    maas_sonuc = None
    arac_sonuc = None
    active_tab = request.args.get("tab", "maas")
    lead_status = request.args.get("lead")
    lead_source = request.args.get("source")

    site_url = request.url_root.rstrip("/") + "/"
    page_title = "Maaş Gelir Gider ve Araç Kullanım Maliyeti Hesaplama | FinansKral"
    meta_description = (
        "Maaş gelir gider hesaplama, net maaş analizi, araç kullanım maliyeti ve "
        "yakıt gideri hesaplama aracı. Türkiye için pratik finans hesaplama platformu."
    )

    fuel_prices = {
        "istanbul_avrupa": {
            "label": "İstanbul Avrupa",
            "gasoline": 62.70,
            "diesel": 71.59,
            "lpg": 34.99,
        },
        "ankara": {
            "label": "Ankara",
            "gasoline": 63.67,
            "diesel": 72.71,
            "lpg": 34.87,
        },
        "izmir": {
            "label": "İzmir",
            "gasoline": 63.94,
            "diesel": 72.99,
            "lpg": 34.79,
        },
    }

    selected_city = request.form.get("fuel_city", "istanbul_avrupa")
    if selected_city not in fuel_prices:
        selected_city = "istanbul_avrupa"

    selected_fuel_data = fuel_prices[selected_city]
    fuel_note = "Yakıt fiyatları şehir ve istasyona göre değişebilir."

    maas_form = {
        "brut": "",
        "yemek": "",
        "yol": "",
        "kira": "",
        "fatura": "",
        "market": "",
        "diger": "",
    }

    arac_form = {
        "fuel_city": selected_city,
        "km": "",
        "tuketim": "",
        "yakit_tipi": "gasoline",
        "yakit": "",
        "sigorta": "",
        "bakim": "",
        "mtv": "",
        "car_income": "",
        "manuel_yakit": "off",
    }

    if request.method == "POST":
        form_tipi = request.form.get("form_type")

        if form_tipi == "maas":
            active_tab = "maas"

            maas_form["brut"] = request.form.get("brut", "")
            maas_form["yemek"] = request.form.get("yemek", "")
            maas_form["yol"] = request.form.get("yol", "")
            maas_form["kira"] = request.form.get("kira", "")
            maas_form["fatura"] = request.form.get("fatura", "")
            maas_form["market"] = request.form.get("market", "")
            maas_form["diger"] = request.form.get("diger", "")

            brut = safe_float(request.form.get("brut"))
            yemek = safe_float(request.form.get("yemek"))
            yol = safe_float(request.form.get("yol"))
            kira = safe_float(request.form.get("kira"))
            fatura = safe_float(request.form.get("fatura"))
            market = safe_float(request.form.get("market"))
            diger = safe_float(request.form.get("diger"))

            sgk = brut * 0.14
            issizlik = brut * 0.01
            vergi_matrahi = brut - sgk - issizlik
            vergi = vergi_matrahi * 0.15

            toplam_kesinti = sgk + issizlik + vergi
            yan_haklar = yemek + yol
            net_maas = brut - toplam_kesinti + yan_haklar
            isveren_sgk = brut * 0.155
            isveren_maliyeti = brut + isveren_sgk + yemek + yol

            total_expenses = {
                "Kira": kira,
                "Fatura": fatura,
                "Market": market,
                "Diğer": diger,
            }
            toplam_gider = sum(total_expenses.values())
            ay_sonu_kalan = net_maas - toplam_gider
            yillik_gider = toplam_gider * 12
            yillik_kalan = ay_sonu_kalan * 12
            savings_rate = (ay_sonu_kalan / net_maas * 100) if net_maas > 0 else 0
            gider_orani = (toplam_gider / net_maas * 100) if net_maas > 0 else 100
            maas_skor = round(clamp(100 - gider_orani, 0, 100))
            biggest_expense_label, biggest_expense_value = max(total_expenses.items(), key=lambda item: item[1])
            benchmark_gap = ay_sonu_kalan - (net_maas * (AVERAGE_SAVINGS_RATE / 100))

            if maas_skor >= 70:
                maas_skor_yorum = "Güçlü denge"
            elif maas_skor >= 40:
                maas_skor_yorum = "Orta denge"
            else:
                maas_skor_yorum = "Zayıf denge"

            if benchmark_gap >= 0:
                benchmark_text = "Ortalama birikim seviyesinin üzerindesin"
            else:
                benchmark_text = "Ortalama birikim seviyesinin altındasın"

            salary_tips = build_salary_tips(savings_rate, biggest_expense_label, ay_sonu_kalan)

            salary_viral = build_viral_salary_lines(net_maas, toplam_gider, ay_sonu_kalan, biggest_expense_label)

            maas_sonuc = {
                "brut_maas": format_tl(brut),
                "sgk": format_tl(sgk),
                "issizlik": format_tl(issizlik),
                "vergi_matrahi": format_tl(vergi_matrahi),
                "vergi": format_tl(vergi),
                "toplam_kesinti": format_tl(toplam_kesinti),
                "yan_haklar": format_tl(yan_haklar),
                "maastan_kalan": format_tl(brut - toplam_kesinti),
                "isveren_sgk": format_tl(isveren_sgk),
                "isveren_maliyeti": format_tl(isveren_maliyeti),
                "net_maas": format_tl(net_maas),
                "kira": format_tl(kira),
                "fatura": format_tl(fatura),
                "market": format_tl(market),
                "diger": format_tl(diger),
                "toplam_gider": format_tl(toplam_gider),
                "ay_sonu_kalan": format_tl(ay_sonu_kalan),
                "ay_sonu_kalan_raw": ay_sonu_kalan,
                "net_maas_raw": net_maas,
                "toplam_gider_raw": toplam_gider,
                "maas_skor": maas_skor,
                "maas_skor_yorum": maas_skor_yorum,
                "yillik_gider": format_tl(yillik_gider),
                "yillik_kalan": format_tl(yillik_kalan),
                "savings_rate": round(savings_rate, 1),
                "biggest_expense_label": biggest_expense_label,
                "biggest_expense_value": format_tl(biggest_expense_value),
                "biggest_expense_raw": biggest_expense_value,
                "benchmark_gap": format_tl(benchmark_gap),
                "benchmark_gap_raw": benchmark_gap,
                "benchmark_text": benchmark_text,
                "tips": salary_tips,
                "expense_labels": list(total_expenses.keys()),
                "expense_values": list(total_expenses.values()),
                "viral": salary_viral,
            }

        elif form_tipi == "arac":
            active_tab = "arac"

            arac_form["fuel_city"] = request.form.get("fuel_city", "istanbul_avrupa")
            arac_form["km"] = request.form.get("km", "")
            arac_form["tuketim"] = request.form.get("tuketim", "")
            arac_form["yakit_tipi"] = request.form.get("yakit_tipi", "gasoline")
            arac_form["yakit"] = request.form.get("yakit", "")
            arac_form["sigorta"] = request.form.get("sigorta", "")
            arac_form["bakim"] = request.form.get("bakim", "")
            arac_form["mtv"] = request.form.get("mtv", "")
            arac_form["car_income"] = request.form.get("car_income", "")
            arac_form["manuel_yakit"] = request.form.get("manuel_yakit", "off")

            selected_city = arac_form["fuel_city"]
            if selected_city not in fuel_prices:
                selected_city = "istanbul_avrupa"
                arac_form["fuel_city"] = selected_city

            selected_fuel_data = fuel_prices[selected_city]

            km = safe_float(request.form.get("km"))
            tuketim = safe_float(request.form.get("tuketim"))
            yakit = safe_float(request.form.get("yakit"))
            sigorta = safe_float(request.form.get("sigorta"))
            bakim = safe_float(request.form.get("bakim"))
            mtv = safe_float(request.form.get("mtv"))
            car_income = safe_float(request.form.get("car_income"))

            aylik_yakit = (km / 100) * tuketim * yakit
            aylik_sigorta = sigorta / 12
            aylik_bakim = bakim / 12
            aylik_mtv = mtv / 12
            yearly_total = sigorta + bakim + mtv + (aylik_yakit * 12)
            aylik_toplam = aylik_yakit + aylik_sigorta + aylik_bakim + aylik_mtv
            km_basi_maliyet = aylik_toplam / km if km > 0 else 0
            arac_skor = round(clamp(100 - (km_basi_maliyet * 10), 0, 100))
            vehicle_costs = {
                "Yakıt": aylik_yakit,
                "Sigorta": aylik_sigorta,
                "Bakım": aylik_bakim,
                "MTV": aylik_mtv,
            }
            biggest_car_cost_label, biggest_car_cost_value = max(vehicle_costs.items(), key=lambda item: item[1])
            yakit_share_pct = (aylik_yakit / aylik_toplam * 100) if aylik_toplam > 0 else 0
            benchmark_diff = aylik_toplam - AVERAGE_MONTHLY_CAR_COST
            cheaper_if_reduce_km = aylik_toplam - (((km * 0.85) / 100) * tuketim * yakit + aylik_sigorta + aylik_bakim + aylik_mtv)
            cheaper_if_better_efficiency = aylik_toplam - ((km / 100) * (tuketim * 0.8) * yakit + aylik_sigorta + aylik_bakim + aylik_mtv)

            if arac_skor >= 70:
                arac_skor_yorum = "Verimli kullanım"
            elif arac_skor >= 40:
                arac_skor_yorum = "Orta maliyet"
            else:
                arac_skor_yorum = "Yüksek maliyet"

            if benchmark_diff <= 0:
                benchmark_text = "Türkiye ortalama araç giderinin altında görünüyorsun"
            else:
                benchmark_text = "Türkiye ortalama araç giderinin üzerindesin"

            car_tips = build_car_tips(km_basi_maliyet, yakit_share_pct, arac_skor)
            affordability = build_car_affordability(car_income, aylik_toplam)
            car_viral = build_viral_car_lines(aylik_toplam, yearly_total, km_basi_maliyet, affordability)

            arac_sonuc = {
                "aylik_yakit": format_tl(aylik_yakit),
                "aylik_sigorta": format_tl(aylik_sigorta),
                "aylik_bakim": format_tl(aylik_bakim),
                "aylik_mtv": format_tl(aylik_mtv),
                "aylik_toplam": format_tl(aylik_toplam),
                "km_basi_maliyet": format_tl(km_basi_maliyet),
                "aylik_yakit_raw": aylik_yakit,
                "aylik_sigorta_raw": aylik_sigorta,
                "aylik_bakim_raw": aylik_bakim,
                "aylik_mtv_raw": aylik_mtv,
                "aylik_toplam_raw": aylik_toplam,
                "arac_skor": arac_skor,
                "arac_skor_yorum": arac_skor_yorum,
                "yearly_total": format_tl(yearly_total),
                "benchmark_diff": format_tl(benchmark_diff),
                "benchmark_diff_raw": benchmark_diff,
                "benchmark_text": benchmark_text,
                "biggest_cost_label": biggest_car_cost_label,
                "biggest_cost_value": format_tl(biggest_car_cost_value),
                "tips": car_tips,
                "save_reduce_km": format_tl(cheaper_if_reduce_km),
                "save_better_efficiency": format_tl(cheaper_if_better_efficiency),
                "cost_labels": list(vehicle_costs.keys()),
                "cost_values": list(vehicle_costs.values()),
                "affordability": affordability,
                "car_income": format_tl(car_income) if car_income > 0 else "Gelir girilmedi",
                "viral": car_viral,
            }

    news_items = get_google_news_items(limit=12)
    crypto_items = get_crypto_prices(limit=6)

    return render_template(
        "index2.html",
        maas_sonuc=maas_sonuc,
        arac_sonuc=arac_sonuc,
        maas_form=maas_form,
        arac_form=arac_form,
        active_tab=active_tab,
        fuel_prices=fuel_prices,
        fuel_note=fuel_note,
        selected_city=selected_city,
        selected_fuel_data=selected_fuel_data,
        adsense_enabled=ADSENSE_ENABLED,
        adsense_client=ADSENSE_CLIENT,
        adsense_top_slot=ADSENSE_TOP_SLOT,
        adsense_content_slot=ADSENSE_CONTENT_SLOT,
        adsense_inline_slot=ADSENSE_INLINE_SLOT,
        adsense_bottom_slot=ADSENSE_BOTTOM_SLOT,
        adsense_left_slot=ADSENSE_LEFT_SLOT,
        adsense_right_slot=ADSENSE_RIGHT_SLOT,
        site_url=site_url,
        page_title=page_title,
        meta_description=meta_description,
        lead_status=lead_status,
        lead_source=lead_source,
        economy_headlines=ECONOMY_HEADLINES,
        news_items=news_items,
        city_living_costs=CITY_LIVING_COSTS,
        daily_economy_impacts=DAILY_ECONOMY_IMPACTS,
        financial_death_zones=FINANCIAL_DEATH_ZONES,
        crypto_items=crypto_items,
    )


ALLOWED_OFFICE_EXTENSIONS = {
    "excel": {"xlsx", "xlsm"},
    "word": {"docx"},
    "pdf": {"pdf"},
}


def _allowed_file(filename, group):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_OFFICE_EXTENSIONS.get(group, set())


def _safe_upload_name(file_storage, fallback="dosya"):
    filename = secure_filename(file_storage.filename or fallback)
    return filename or fallback


def _build_simple_pdf(title, sections):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("FinansKralTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, spaceAfter=16)
    heading_style = ParagraphStyle("FinansKralHeading", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=16, textColor=colors.HexColor("#0f766e"))
    body_style = ParagraphStyle("FinansKralBody", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=13)

    story = [Paragraph(title, title_style)]
    story.append(Paragraph("FinansKral Ofis Araclari ile olusturuldu.", body_style))
    story.append(Spacer(1, 12))

    for section in sections:
        story.append(Paragraph(section.get("heading", "Bolum"), heading_style))
        rows = section.get("rows")
        if rows:
            clean_rows = []
            for row in rows[:80]:
                clean_rows.append([Paragraph(str(cell)[:160] if cell is not None else "", body_style) for cell in row[:8]])
            if clean_rows:
                col_count = max(len(row) for row in clean_rows)
                for row in clean_rows:
                    while len(row) < col_count:
                        row.append(Paragraph("", body_style))
                table = Table(clean_rows, repeatRows=1 if len(clean_rows) > 1 else 0)
                table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d8dee9")),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eefdf8")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))
                story.append(table)
        else:
            for paragraph in section.get("paragraphs", []):
                story.append(Paragraph(str(paragraph), body_style))
        story.append(Spacer(1, 12))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _excel_to_pdf(file_storage):
    from openpyxl import load_workbook
    filename = _safe_upload_name(file_storage, "excel.xlsx")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(filename).suffix) as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name
    try:
        workbook = load_workbook(tmp_path, data_only=True, read_only=True)
        sections = []
        for sheet in workbook.worksheets[:6]:
            rows = []
            for row in sheet.iter_rows(max_row=45, max_col=8, values_only=True):
                if any(cell not in (None, "") for cell in row):
                    rows.append(["" if cell is None else cell for cell in row])
            if rows:
                sections.append({"heading": sheet.title, "rows": rows})
        if not sections:
            sections = [{"heading": "Excel Ozeti", "paragraphs": ["Dosyada goruntulenecek veri bulunamadi."]}]
        return _build_simple_pdf("Excel PDF Donusumu", sections), filename.rsplit(".", 1)[0] + ".pdf"
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def _word_to_pdf(file_storage):
    from docx import Document
    filename = _safe_upload_name(file_storage, "word.docx")
    with tempfile.NamedTemporaryFile(delete=False, suffix=Path(filename).suffix) as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name
    try:
        doc = Document(tmp_path)
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        if not paragraphs:
            paragraphs = ["Dosyada goruntulenecek metin bulunamadi."]
        sections = [{"heading": "Word Belgesi", "paragraphs": paragraphs[:140]}]
        return _build_simple_pdf("Word PDF Donusumu", sections), filename.rsplit(".", 1)[0] + ".pdf"
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def _merge_pdfs(files):
    from pypdf import PdfReader, PdfWriter
    writer = PdfWriter()
    used = 0
    temp_paths = []
    try:
        for file_storage in files:
            if not file_storage or not file_storage.filename or not _allowed_file(file_storage.filename, "pdf"):
                continue
            filename = _safe_upload_name(file_storage, "dosya.pdf")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                file_storage.save(tmp.name)
                temp_paths.append(tmp.name)
            reader = PdfReader(tmp.name)
            for page in reader.pages:
                writer.add_page(page)
            used += 1
        if used == 0:
            raise ValueError("En az bir PDF dosyasi yukleyin.")
        buffer = io.BytesIO()
        writer.write(buffer)
        buffer.seek(0)
        return buffer, "finanskral-birlestirilmis.pdf"
    finally:
        for path in temp_paths:
            try:
                os.remove(path)
            except OSError:
                pass


def _compress_pdf(file_storage):
    from pypdf import PdfReader, PdfWriter
    filename = _safe_upload_name(file_storage, "dosya.pdf")
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        file_storage.save(tmp.name)
        tmp_path = tmp.name
    try:
        reader = PdfReader(tmp_path)
        writer = PdfWriter()
        for page in reader.pages:
            try:
                page.compress_content_streams()
            except Exception:
                pass
            writer.add_page(page)
        buffer = io.BytesIO()
        writer.write(buffer)
        buffer.seek(0)
        return buffer, filename.rsplit(".", 1)[0] + "-sikistirilmis.pdf"
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


@app.route("/capture-lead", methods=["POST"])
def capture_lead():
    source = request.form.get("source", "maas")
    save_lead({
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "name": request.form.get("lead_name", "").strip(),
        "email": request.form.get("lead_email", "").strip(),
        "phone": request.form.get("lead_phone", "").strip(),
        "goal": request.form.get("lead_goal", "").strip(),
        "summary": request.form.get("lead_summary", "").strip(),
        "page": request.url_root.rstrip("/") + "/",
    })
    return redirect(url_for("home", lead="success", source=source, tab=source))

@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/privacy")
def privacy():
    return render_template("privacy.html")


@app.route("/contact")
def contact():
    return render_template("contact.html")



@app.route("/cookies")
def cookies():
    return render_template("cookies.html")


@app.route("/terms")
def terms():
    return render_template("terms.html")



@app.route("/haberler")
def news_page():
    site_url = request.url_root.rstrip("/") + "/haberler"
    page_title = "Güncel Ekonomi Haberleri | FinansKral"
    meta_description = "Dolar, euro, akaryakıt, enflasyon, maaş ve bütçe gündemini Google News RSS akışıyla tek ekranda takip edin."
    news_items = get_google_news_items(limit=24)
    return render_template(
        "news.html",
        news_items=news_items,
        site_url=site_url,
        page_title=page_title,
        meta_description=meta_description,
    )


@app.route("/ofis-araclari")
def office_tools():
    site_url = request.url_root.rstrip("/") + "/ofis-araclari"
    return render_template(
        "office_tools.html",
        site_url=site_url,
        page_title="Ofis Araçları | Excel PDF, Word PDF, PDF Birleştir | FinansKral",
        meta_description="FinansKral Ofis Araçları ile Excel PDF, Word PDF, PDF birleştirme, PDF sıkıştırma, maaş zam, fazla mesai ve enflasyon kaybı hesaplama işlemlerini tek ekranda yapın.",
    )


@app.route("/ofis-araclari/excel-pdf", methods=["POST"])
def excel_pdf_tool():
    file_storage = request.files.get("excel_file")
    if not file_storage or not _allowed_file(file_storage.filename, "excel"):
        return redirect(url_for("office_tools", error="excel"))
    pdf_buffer, output_name = _excel_to_pdf(file_storage)
    return send_file(pdf_buffer, mimetype="application/pdf", as_attachment=True, download_name=output_name)


@app.route("/ofis-araclari/word-pdf", methods=["POST"])
def word_pdf_tool():
    file_storage = request.files.get("word_file")
    if not file_storage or not _allowed_file(file_storage.filename, "word"):
        return redirect(url_for("office_tools", error="word"))
    pdf_buffer, output_name = _word_to_pdf(file_storage)
    return send_file(pdf_buffer, mimetype="application/pdf", as_attachment=True, download_name=output_name)


@app.route("/ofis-araclari/pdf-birlestir", methods=["POST"])
def merge_pdf_tool():
    files = request.files.getlist("pdf_files")
    try:
        pdf_buffer, output_name = _merge_pdfs(files)
    except Exception:
        return redirect(url_for("office_tools", error="merge"))
    return send_file(pdf_buffer, mimetype="application/pdf", as_attachment=True, download_name=output_name)


@app.route("/ofis-araclari/pdf-sikistir", methods=["POST"])
def compress_pdf_tool():
    file_storage = request.files.get("pdf_file")
    if not file_storage or not _allowed_file(file_storage.filename, "pdf"):
        return redirect(url_for("office_tools", error="compress"))
    pdf_buffer, output_name = _compress_pdf(file_storage)
    return send_file(pdf_buffer, mimetype="application/pdf", as_attachment=True, download_name=output_name)

@app.route("/blog")
def blog():
    site_url = request.url_root.rstrip("/") + "/blog"
    return render_template("blog.html", posts=BLOG_POSTS, site_url=site_url)


@app.route("/blog/<slug>")
def blog_detail(slug):
    post = next((item for item in BLOG_POSTS if item["slug"] == slug), None)
    if post is None:
        return redirect(url_for("blog"))
    site_url = request.url_root.rstrip("/") + url_for("blog_detail", slug=slug)
    return render_template("blog_detail.html", post=post, posts=BLOG_POSTS, site_url=site_url)


if __name__ == "__main__":
    app.run(debug=True)
